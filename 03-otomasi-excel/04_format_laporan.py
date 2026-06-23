#!/usr/bin/env python3
"""04 — Format Laporan Excel Otomatis
Contoh: buat laporan yang sudah diformat (header, border, warna, lebar kolom).
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Siapkan data
data = {
    "No": [1, 2, 3, 4],
    "NIP": ["198501152010011004", "198803122012012003", "199006012015011001", "199201252018012005"],
    "Nama": ["Budi Santoso", "Siti Rahayu", "Ahmad Hidayat", "Dewi Lestari"],
    "Golongan": ["III/c", "IV/a", "III/b", "III/a"],
    "Gaji Pokok": [5_500_000, 6_200_000, 5_000_000, 4_500_000],
}

df = pd.DataFrame(data)
file_out = "laporan_formatted.xlsx"
df.to_excel(file_out, index=False, startrow=1, sheet_name="Rekap")

# Buka workbook untuk diformat
wb = load_workbook(file_out)
ws = wb.active

# Tambah judul
ws.merge_cells("A1:E1")
ws["A1"] = "REKAP DATA PEGAWAI — JUNI 2026"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center")

# Format header
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

for cell in ws[2]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Format semua cell (border + alignment)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        if isinstance(cell.value, int) and cell.value > 1000:
            cell.number_format = "#,##0"

# Lebar kolom
widths = {"A": 6, "B": 22, "C": 18, "D": 12, "E": 15}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(file_out)
print(f"✅ Laporan terformat tersimpan: {file_out}")
